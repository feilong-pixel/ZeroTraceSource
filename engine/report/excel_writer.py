"""Excel report output for investigation results."""

from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path

from engine.models import RESULT_HEADERS, SearchRequest, SearchResult

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - reported when writer is called.
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


def write_excel_report(request: SearchRequest, results: list[SearchResult]) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for report output") from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    result_sheet = workbook.create_sheet("SearchResults")
    keyword_sheet = workbook.create_sheet("Keywords")
    root_sheet = workbook.create_sheet("SearchRoots")

    write_summary(summary, request, results)
    write_results(result_sheet, results)
    write_keywords(keyword_sheet, request.keywords)
    write_roots(root_sheet, request.roots)

    for sheet in workbook.worksheets:
        style_sheet(sheet)
        autofit(sheet, get_column_letter)

    request.output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(request.output_path)
    workbook.close()
    return request.output_path


def write_summary(sheet, request: SearchRequest, results: list[SearchResult]) -> None:
    counts = Counter(result.category for result in results)
    rows = [
        ("Investigation title", request.title),
        ("Created time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Search roots", len(request.roots)),
        ("Keyword count", len(request.keywords)),
        ("Total hit count", len(results)),
        ("Text hit count", counts.get("Text", 0)),
        ("Excel hit count", counts.get("Excel", 0)),
        ("Error count", counts.get("Error", 0)),
    ]

    sheet.append(["Item", "Value"])
    for row in rows:
        sheet.append(row)


def write_results(sheet, results: list[SearchResult]) -> None:
    sheet.append(RESULT_HEADERS)
    for index, result in enumerate(results, start=1):
        sheet.append(
            [
                index,
                result.investigation_title,
                result.keyword,
                result.category,
                str(result.file_path),
                result.file_name,
                result.extension,
                result.sheet_name,
                result.cell_address,
                result.line_number,
                result.column_number,
                result.encoding,
                result.matched_text,
                result.before_text,
                result.after_text,
                result.confirm_status,
                result.notes,
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_keywords(sheet, keywords: list[str]) -> None:
    sheet.append(["No", "Keyword"])
    for index, keyword in enumerate(keywords, start=1):
        sheet.append([index, keyword])


def write_roots(sheet, roots: list[Path]) -> None:
    sheet.append(["No", "SearchRoot"])
    for index, root in enumerate(roots, start=1):
        sheet.append([index, str(root)])


def style_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def autofit(sheet, get_column_letter) -> None:
    max_width = 72
    for column in sheet.columns:
        column_letter = get_column_letter(column[0].column)
        width = 10
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        sheet.column_dimensions[column_letter].width = width
